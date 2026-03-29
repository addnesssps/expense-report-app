import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers


def generate_excel_report(expenses: list[dict], applicant: str = "",
                          company: str = "", period_start: str = "",
                          period_end: str = "") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "経費精算書"

    # スタイル定義
    title_font = Font(name="游ゴシック", size=16, bold=True)
    header_font = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
    cell_font = Font(name="游ゴシック", size=10)
    info_font = Font(name="游ゴシック", size=11)
    total_font = Font(name="游ゴシック", size=11, bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # 列幅
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 30

    # タイトル
    ws.merge_cells("A1:F1")
    ws["A1"] = "経費精算書"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # 情報欄
    row = 3
    ws[f"A{row}"] = "会社名："
    ws[f"A{row}"].font = info_font
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"B{row}"] = company
    ws[f"B{row}"].font = info_font

    ws[f"D{row}"] = "申請者："
    ws[f"D{row}"].font = info_font
    ws.merge_cells(f"E{row}:F{row}")
    ws[f"E{row}"] = applicant
    ws[f"E{row}"].font = info_font

    row = 4
    ws[f"A{row}"] = "対象期間："
    ws[f"A{row}"].font = info_font
    ws.merge_cells(f"B{row}:C{row}")
    period_text = ""
    if period_start and period_end:
        period_text = f"{period_start} 〜 {period_end}"
    elif period_start:
        period_text = f"{period_start} 〜"
    ws[f"B{row}"] = period_text
    ws[f"B{row}"].font = info_font

    ws[f"D{row}"] = "作成日："
    ws[f"D{row}"].font = info_font
    ws.merge_cells(f"E{row}:F{row}")
    ws[f"E{row}"] = datetime.now().strftime("%Y-%m-%d")
    ws[f"E{row}"].font = info_font

    # ヘッダー行
    row = 6
    headers = ["No.", "日付", "支払先", "科目", "金額", "摘要"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # データ行
    total_amount = 0
    for i, exp in enumerate(expenses, 1):
        row += 1
        amount = exp.get("amount", 0)
        total_amount += amount

        values = [
            i,
            exp.get("date", ""),
            exp.get("vendor", ""),
            exp.get("category", ""),
            amount,
            exp.get("description", ""),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = center
            elif col_idx == 5:
                cell.alignment = right_align
                cell.number_format = '#,##0"円"'

    # 合計行
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "合計"
    ws[f"A{row}"].font = total_font
    ws[f"A{row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"A{row}"].fill = total_fill
    ws[f"A{row}"].border = thin_border
    for col in range(2, 5):
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).fill = total_fill

    total_cell = ws.cell(row=row, column=5, value=total_amount)
    total_cell.font = total_font
    total_cell.alignment = right_align
    total_cell.border = thin_border
    total_cell.fill = total_fill
    total_cell.number_format = '#,##0"円"'

    ws.cell(row=row, column=6).border = thin_border
    ws.cell(row=row, column=6).fill = total_fill

    # 科目別集計
    row += 2
    ws[f"A{row}"] = "【科目別集計】"
    ws[f"A{row}"].font = Font(name="游ゴシック", size=11, bold=True)

    category_totals = {}
    for exp in expenses:
        cat = exp.get("category", "その他")
        category_totals[cat] = category_totals.get(cat, 0) + exp.get("amount", 0)

    row += 1
    for col_idx, header in enumerate(["科目", "件数", "金額"], 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    for cat, total in sorted(category_totals.items()):
        row += 1
        count = sum(1 for e in expenses if e.get("category") == cat)
        ws.cell(row=row, column=1, value=cat).font = cell_font
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=count).font = cell_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = center
        amount_cell = ws.cell(row=row, column=3, value=total)
        amount_cell.font = cell_font
        amount_cell.border = thin_border
        amount_cell.alignment = right_align
        amount_cell.number_format = '#,##0"円"'

    # BytesIOに保存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
